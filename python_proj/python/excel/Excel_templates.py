import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font

def create_excel_template(template_path, sheet_name, header, data=None):
    # Create a new workbook
    wb = openpyxl.Workbook()

    # Get the active sheet
    sheet = wb.active
    sheet.title = sheet_name

    # Add header to the sheet
    for col_num, col_title in enumerate(header, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = col_title
        cell.font = Font(bold=True)

    # Add data to the sheet, if provided
    if data:
        for row_num, row_data in enumerate(data, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value

    # Save the template
    wb.save(template_path)
    print(f"Excel template created and saved at: {template_path}")

def update_excel_template(template_path, new_data):
    # Load the existing workbook
    wb = load_workbook(template_path)

    # Get the active sheet
    sheet = wb.active

    # Find the next empty row
    next_row = sheet.max_row + 1

    # Add new data to the sheet
    for col_num, value in enumerate(new_data, 1):
        cell = sheet.cell(row=next_row, column=col_num)
        cell.value = value

    # Save the updated template
    wb.save(template_path)
    print(f"Excel template updated and saved at: {template_path}")

# Example usage:
template_path = 'C:/python_proj/python/excel/excel_template.xlsx'
template_sheet_name = 'Sheet1'
template_header = ['Name', 'Age', 'City']

# Create a template with headers
create_excel_template(template_path, template_sheet_name, template_header)

# Update the template with new data
new_data_row = ['John Doe', 30, 'New York']
update_excel_template(template_path, new_data_row)
