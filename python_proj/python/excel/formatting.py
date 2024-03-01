from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from copy import copy

def copy_conditional_formatting(source_ws, source_range, target_ws, target_range):
    source_cells = source_ws[source_range]
    target_cells = target_ws[target_range]

    for source_row, target_row in zip(source_cells, target_cells):
        for source_cell, target_cell in zip(source_row, target_row):
            # Copy conditional formatting rules
            if source_cell.has_style:
                target_cell.fill = copy(source_cell.fill)

# Example usage:
source_excel_file = 'C:/python_proj/python/excel/workbook.xlsx'
target_excel_file = 'C:/pto/workbook.xlsx'

# Open workbooks
source_wb = load_workbook(source_excel_file)
target_wb = load_workbook(target_excel_file)

# Specify sheet names and cell ranges
source_sheet_name = 'Sheet1'
target_sheet_name = 'Sheet2'
source_range = 'A1:A10'
target_range = 'B1:B10'

# Get source and target worksheets
source_ws = source_wb[source_sheet_name]
target_ws = target_wb[target_sheet_name]

# Copy conditional formatting rules
copy_conditional_formatting(source_ws, source_range, target_ws, target_range)

# Save the modified target workbook
target_wb.save('C:/pto/workbook_with_formatting.xlsx')
